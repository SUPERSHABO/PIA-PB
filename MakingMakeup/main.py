import aritmetics
import json
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import requests
from tabulate import tabulate
import top_worst_products as twp

def hacer_solicitud_api():
    try:
        response = requests.get("http://makeup-api.herokuapp.com/api/v1/products.json")
        response.raise_for_status()  # Manejar posibles errores HTTP
        return json.loads(response.text)
    except requests.RequestException as e:
        print(f"Error en la solicitud API: {e}")
        return []

def listar_productos_maquillaje(data):
    if data:
        headers = ["id", "brand", "name", "price"]
        filas = [[producto["id"], producto["brand"], producto["name"], producto["price"]]
                 for producto in data]
        print(tabulate(filas, headers=headers, tablefmt="grid"))

def obtener_precios_maquillaje(data):
    precios = [float(product["price"]) for product in data if product.get("price") is not None and product.get("price") != ""]
    return precios

def graficar_precios_maquillaje(precios):
    if precios:  
        plt.plot(range(1, len(precios) + 1), precios)
        plt.xlabel("Producto")
        plt.ylabel("Precio")
        plt.title("Precios de Productos de Maquillaje")
        plt.show()
    else:
        print("No hay datos válidos de precios para graficar.")

def filtrar_productos_por_marca(data, marca):
    if marca is not None and marca != "":
        return [product for product in data if product.get('brand') is not None and str(product.get('brand')).lower() == marca.lower() and product.get('price') is not None and float(product.get('price')) != 0]
    else:
        print("La marca ingresada no es válida.")
        return []

def guardar_en_txt(datos, nombre_archivo_txt, marca_filtrada=None):
    with open(nombre_archivo_txt, 'w') as archivo_txt:
        for product in datos:
            if marca_filtrada is None or (product.get('brand') is not None and str(product.get('brand')).lower() == marca_filtrada.lower() and product.get('price') is not None and float(product.get('price')) != 0):
                archivo_txt.write(f"Producto: {product['name']}\n")
                archivo_txt.write(f"Marca: {product['brand']}\n")
                archivo_txt.write(f"Categoría: {product['category']}\n")
                archivo_txt.write(f"Precio: {product['price']}\n")
                archivo_txt.write("-------------------------\n")

def guardar_en_excel(datos, nombre_archivo_excel):
    libro_excel = openpyxl.Workbook()
    hoja_excel = libro_excel.active

    # Escribir encabezados en la primera fila
    encabezados = ["Producto", "Marca", "Categoría", "Precio"]
    for columna, encabezado in enumerate(encabezados, start=1):
        hoja_excel.cell(row=1, column=columna, value=encabezado)

    # Escribir datos en las filas siguientes
    for fila, product in enumerate(datos, start=2):
        hoja_excel.cell(row=fila, column=1, value=product.get('name', ''))
        hoja_excel.cell(row=fila, column=2, value=product.get('brand', ''))
        hoja_excel.cell(row=fila, column=3, value=product.get('category', ''))
        hoja_excel.cell(row=fila, column=4, value=product.get('price', ''))

    # Guardar el archivo Excel
    libro_excel.save(nombre_archivo_excel)

def graficar_datos_excel(nombre_archivo_excel):
    try:
        libro_excel = openpyxl.load_workbook(nombre_archivo_excel)
        hoja_excel = libro_excel.active

        precios = []

        for fila in hoja_excel.iter_rows(min_row=2, max_row=hoja_excel.max_row, min_col=4, max_col=4):
            for celda in fila:
                if celda.value is not None:
                    precios.append(celda.value)

        if precios:
            plt.plot(range(1, len(precios) + 1), precios)
            plt.xlabel("Producto")
            plt.ylabel("Precio")
            plt.title("Precios de Productos de Maquillaje desde Excel")
            plt.show()
        else:
            print("No hay datos válidos de precios en el archivo Excel para graficar.")
    except Exception as e:
        print(f"Error al cargar el archivo Excel: {e}")

def consultar_productos_por_categoria(data, categoria):
    productos_categoria = [product for product in data if product.get('category') == categoria]
    return productos_categoria

def guardar_en_txt_por_categoria(datos, nombre_archivo_txt, categoria):
    with open(nombre_archivo_txt, 'w') as archivo_txt:
        for product in datos:
            if product.get('category') == categoria:
                archivo_txt.write(f"Producto: {product['name']}\n")
                archivo_txt.write(f"Marca: {product['brand']}\n")
                archivo_txt.write(f"Categoría: {product['category']}\n")
                archivo_txt.write(f"Precio: {product['price']}\n")
                archivo_txt.write("-------------------------\n")

def graficar_datos_txt(nombre_archivo_txt):
  try:
      with open(nombre_archivo_txt, 'r') as archivo_txt:
          precios = []
          for linea in archivo_txt:
              if "Precio:" in linea:
                  precio_str = linea.split(":")[1].strip()
                  if precio_str.lower() != 'none':  # Verificar si el valor no es 'None'
                      precio = float(precio_str)
                      precios.append(precio)

          if precios:
              plt.plot(range(1, len(precios) + 1), precios)
              plt.xlabel("Producto")
              plt.ylabel("Precio")
              plt.title("Precios de Productos de Maquillaje desde TXT")
              plt.show()
          else:
              print("No hay datos válidos de precios en el archivo TXT para graficar.")
  except Exception as e:
      print(f"Error al cargar el archivo TXT: {e}")


def menu_consulta_por_categoria(data):
    categoria = input("Ingrese la categoría de productos que desea consultar: ")
    productos_categoria = consultar_productos_por_categoria(data, categoria)

    if productos_categoria:
        print(f"Productos de la categoría '{categoria}':")
        for product in productos_categoria:
            print(f"Producto: {product['name']}, Marca: {product['brand']}, Precio: {product['price']}")

        # Preguntar si desea guardar en un archivo TXT
        guardar_txt = input("¿Desea guardar estos productos en un archivo TXT? (Sí/No): ").lower()
        if guardar_txt == "si":
            filename_txt = input("Ingrese el nombre del archivo TXT: ")
            guardar_en_txt_por_categoria(productos_categoria, filename_txt, categoria)
            print(f"Datos guardados en '{filename_txt}'.")

            # Preguntar si desea graficar los datos guardados
            graficar_guardados = input("¿Desea graficar los datos guardados? (Sí/No): ").lower()
            if graficar_guardados == "si":
                graficar_datos_txt(filename_txt)
    else:
        print(f"No hay productos en la categoría '{categoria}'.")

def graficar_tendencia(precios):
  if precios:
    x = np.arange(1, len (precios) + 1)
    y = np.array(precios)

    #ajuste polinomico 
    coeficientes = np.polyfit(x, y, 1)
    polinomio = np.poly1d(coeficientes)
    tendencia = polinomio(x)

    #graficar puntos de datos
    plt.scatter(x, y, label = "Datos")

    #graficar tendencia
    plt.plot(x, tendencia, color = "red", label = "Tendencia")

    #etiquetas y titulos
    plt.xlabel("Producto")
    plt.ylabel("Precio")
    plt.title("Tendencia de precios de productos de maquillaje")
    plt.legend()

    #mostrar grafico  
    plt.show()
  else:
    print("Primero obten precios antes de intentar graficar")

    
def mostrar_menu_principal(data):
    opcion = None
    precios = []  # Define precios aquí para que sea accesible en todo el bucle.
    while opcion != "0":
        print("¡Bienvenido a la búsqueda de productos de maquillaje!")
        print("1 - Obtener datos")
        print("2 - Gráficas")
        print("3 - Exportar")
        print("4 - Estadísticas")
        print("0 - Salir")
        opcion = input("Ingrese su opción: ")

        if opcion == "1":
            opcionI = None
            while opcionI != "0":
                print("Menú de obtener datos")
                print("1 - Listar TODOS los productos de maquillaje")
                print("2 - Obtener precios de productos de maquillaje")
                print("3 - Consultar productos por marca")
                print("4 - Consultar productos por categoría")
                print("0 - Salir del menú de obtener datos")
                opcionI = input("Ingrese su opción: ")

                if opcionI == "1":
                    listar_productos_maquillaje(data)

                elif opcionI == "2":
                    precios = obtener_precios_maquillaje(data)
                    print("Precios de productos de maquillaje:")
                    for precio in precios:
                        print(precio)

                elif opcionI == "3":
                    marca = input("Ingrese la marca para filtrar los productos: ")
                    productos_filtrados = filtrar_productos_por_marca(data, marca)
                    print("Productos filtrados por marca:")
                    for product in productos_filtrados:
                        print(f"Producto: {product['name']}, Marca: {product['brand']}")

                elif opcionI == "4":
                    menu_consulta_por_categoria(data)

                elif opcionI == "0":
                    print("Saliendo del menú de obtener datos...")

                else:
                    print("Opción inválida. Por favor, intente nuevamente.")

        elif opcion == "2":
            opcionII = None
            while opcionII != "0":
                print("Menú de gráficas")
                print("1 - Graficar precios de productos de maquillaje")
                print ("2 - Graficar tendencia de precios")
                print("3 - Graficar datos desde Excel")
                print("4 - Graficar datos desde TXT")
                print("0 - Salir del menú de gráficas")
                opcionII = input("Ingrese su opción: ")

                if opcionII == "1":
                    if precios:
                        graficar_precios_maquillaje(precios)
                        regresar_menu = input("¿Desea volver al menú de gráficas? (Sí/No): ").lower()
                        if regresar_menu != "si":
                            continue
                    else:
                        print("Primero obtén precios antes de intentar graficar.")
                
                elif opcionII == "2":
                  graficar_tendencia(precios)
                
                elif opcionII == "3":
                    filename_excel_grafico = input("Ingrese el nombre del archivo Excel para graficar (sin extensión): ") + ".xlsx"
                    graficar_datos_excel(filename_excel_grafico)

                elif opcionII == "4":
                    filename_txt_grafico = input("Ingrese el nombre del archivo TXT para graficar: ")
                    graficar_datos_txt(filename_txt_grafico)

                elif opcionII == "0":
                    print("Saliendo del menú de gráficas...")

                else:
                    print("Opción inválida. Por favor, intente nuevamente.")

        elif opcion == "3":
            opcionIII = None
            while opcionIII != "0":
                print("Menú de exportar")
                print("1 - Exportar a TXT")
                print("2 - Exportar a Excel")
                print("3 - Filtrar por marca y Exportar a TXT")
                print("0 - Salir del menú de exportar")
                opcionIII = input("Ingrese su opción: ")

                if opcionIII == "1":
                    filename_txt = input("Ingrese el nombre del archivo TXT: ")
                    guardar_en_txt(data, filename_txt)

                elif opcionIII == "2":
                    filename_excel = input("Ingrese el nombre del archivo Excel (sin extensión): ") + ".xlsx"
                    guardar_en_excel(data, filename_excel)

                elif opcionIII == "3":
                    marca_filtrar_txt = input("Ingrese la marca para filtrar los productos en el archivo TXT: ")
                    filename_txt_filtrado = input("Ingrese el nombre del archivo TXT filtrado: ")
                    guardar_en_txt(data, filename_txt_filtrado, marca_filtrar_txt)

                elif opcionIII == "0":
                    print("Saliendo del menú de exportar...")

                else:
                    print("Opción inválida. Por favor, intente nuevamente.")

        elif opcion == "4":
            opcionIV = None
            while opcionIV != "0":
                print("Menú de estadísticas")
                print("1 - Calcular precio promedio")
                print("2 - Calcular precio más alto")
                print("3 - Calcular precio más bajo")
                print("4 - Obtener el producto con mejor rating")
                print("5 - Obtener el producto con el peor rating")
                print("0 - Salir del menú de estadísticas")
                opcionIV = input("Ingrese su opción: ")

                if opcionIV == "1":
                    prom = aritmetics.calcular_precio_promedio(datos_maquillaje)
                    print("Precio promedio de los productos:", prom)
                    print("-------------------------")

                elif opcionIV == "2":
                    name_higher, price_higher = aritmetics.obtener_producto_mas_caro(datos_maquillaje)
                    print("Producto más caro:")
                    print("Nombre:", name_higher)
                    print("Precio:", price_higher)
                    print("-------------------------")

                elif opcionIV == "3":
                    name_lower, price_lower = aritmetics.obtener_producto_mas_barato(datos_maquillaje)
                    print("Producto más caro:")
                    print("Nombre:", name_lower)
                    print("Precio:", price_lower)
                    print("-------------------------")
                
                elif opcionIV == "4":
                    best_n, best_r, best_p = twp.best(datos_maquillaje)
                    print("El producto con mejor rating es:", best_n)
                    print("Rating:", best_r)
                    print("Precio:", best_p)
                
                elif opcionIV == "5":
                    worst_n, worst_r, worst_p = twp.worst(datos_maquillaje)
                    print("El producto con peor rating es:", worst_n)
                    print("Rating:", worst_r)
                    print("Precio:", worst_p)

                elif opcionIV == "0":
                    print("Saliendo del menú de estadísticas...")

                else:
                    print("Opción inválida. Por favor, intente nuevamente.")

        elif opcion == "0":
            print("¡Hasta luego!")

        else:
            print("Opción inválida. Por favor, intente nuevamente.")

# Realiza la solicitud API una vez para evitar solicitudes redundantes.
datos_maquillaje = hacer_solicitud_api()
mostrar_menu_principal(datos_maquillaje)